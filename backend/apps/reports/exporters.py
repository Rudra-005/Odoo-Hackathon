import csv
import io
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse

class BaseExporter:
    def export(self, data, filename="report"):
        raise NotImplementedError

class CSVExporter(BaseExporter):
    def export(self, data, filename="report"):
        if not data:
            return HttpResponse("No data", content_type='text/plain')
            
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        writer = csv.writer(response)
        
        # Write headers
        headers = list(data[0].keys())
        writer.writerow([h.replace('_', ' ').title() for h in headers])
        
        # Write data
        for row in data:
            writer.writerow(list(row.values()))
            
        return response

class ExcelExporter(BaseExporter):
    def export(self, data, filename="report"):
        if not data:
            return HttpResponse("No data", content_type='text/plain')
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report Data"
        
        headers = list(data[0].keys())
        ws.append([h.replace('_', ' ').title() for h in headers])
        
        # Style headers
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            
        for row in data:
            ws.append(list(row.values()))
            
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        wb.save(response)
        return response

class PDFExporter(BaseExporter):
    def export(self, data, filename="report"):
        if not data:
            return HttpResponse("No data", content_type='text/plain')
            
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        title = Paragraph(f"TransitOps - {filename.replace('_', ' ').title()}", styles['Heading1'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        headers = [h.replace('_', ' ').title() for h in data[0].keys()]
        table_data = [headers]
        
        for row in data:
            table_data.append([str(v)[:30] for v in row.values()]) # Truncate long strings for PDF
            
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1'))
        ]))
        
        elements.append(t)
        doc.build(elements)
        
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response

class ExportFactory:
    @staticmethod
    def get_exporter(format_type):
        if format_type == 'csv':
            return CSVExporter()
        elif format_type == 'excel':
            return ExcelExporter()
        elif format_type == 'pdf':
            return PDFExporter()
        else:
            raise ValueError(f"Unsupported format: {format_type}")
